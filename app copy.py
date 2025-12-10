#
import math
import matplotlib
matplotlib.use("Agg")  # 화면 표시 없이 이미지 생성 (macOS 안전)
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import numpy as np
import time
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, redirect, send_file
import threading
from flask_cors import CORS
import io, base64
import os
import webview
import CoolProp.CoolProp as CP
from openpyxl import load_workbook
# from numba import jit
print("THIS IS THE REAL APP.PY")

app = Flask(__name__, static_folder="frontend/dist")
CORS(app)

IS_DEV = os.environ.get("FLASK_ENV") == "development"

NONCONDENSABLES = [
    'Air', 'Nitrogen', 'Oxygen', 'Helium', 'Argon', 'Neon',
    'Hydrogen', 'CO2', 'Methane'
]

def get_phase(fluid, T, P):
    fname = fluid.lower()

    # Noncondensables 우선 처리
    if fluid in NONCONDENSABLES:
        return "Noncondensables"

    # Q(품질) 계산
    try:
        Q = CP.PropsSI('Q', 'T', T, 'P', P, fluid)
    except:
        Q = None

    if fname in ["water", "h2o"]:
        if Q is None:
            try:
                Tsat = CP.PropsSI('T', 'P', P, 'Q', 0, fluid)
                if T < Tsat:
                    return "Water"
                else:
                    return "Vapor"
            except:
                return "Water"

        if Q <= 0:
            return "Water"
        elif 0 < Q < 1:
            return "Steam"
        else:
            return "Vapor"

    if Q is None:
        try:
            Tsat = CP.PropsSI('T', 'P', P, 'Q', 0, fluid)
            if T < Tsat:
                return "Liquid"
            else:
                return "Vapor"
        except:
            return "Vapor"

    if Q <= 0:
        return "Liquid"
    elif 0 < Q < 1:
        return "Steam"
    else:
        return "Vapor"


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve(path):
    # ✔ 개발 환경: React Vite dev server로 redirect
    if IS_DEV:
        return redirect(f"http://localhost:5173/{path}")

    # ✔ 프로덕션 환경: dist 빌드된 파일을 serve
    if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)

    # ✔ SPA: 모든 URL은 index.html
    return send_from_directory(app.static_folder, "index.html")

@app.route("/api/test", methods=["POST"])
def test():
   data = request.get_json()

   temperature_c = data.get('temperature_c')
   pressure_bar = data.get('pressure_bar')
   composition = data.get('composition')

   T_kelvin = temperature_c + 273.15  # K
   P_pascal = pressure_bar * 1e5      

   fluid_string = f"HEOS::{composition}"

   try:
      # 밀도 (Density, kg/m^3)
      rho = CP.PropsSI('D', 'T', T_kelvin, 'P', P_pascal, fluid_string)
        
      # 정압 비열 (Specific Heat at Constant Pressure, J/kg·K)
      cp = CP.PropsSI('Cpmass', 'T', T_kelvin, 'P', P_pascal, fluid_string)
        
      # 열전도도 (Thermal Conductivity, W/m·K)
      k = CP.PropsSI('CONDUCTIVITY', 'T', T_kelvin, 'P', P_pascal, fluid_string)
      
      # 동점성계수 (Dynamic Viscosity, Pa·s)
      mu = CP.PropsSI('VISCOSITY', 'T', T_kelvin, 'P', P_pascal, fluid_string)
      
      # 플란틀 수 (Prandtl Number)
      pr = CP.PropsSI('Prandtl', 'T', T_kelvin, 'P', P_pascal, fluid_string)

      # 4. 파생 물성치 계산
      # 비체적 (Specific Volume, m^3/kg) = 1 / 밀도
      v = 1.0 / rho
      
      # 동점성계수 (Kinematic Viscosity, m^2/s or cSt) = Dynamic Viscosity / Density
      nu = mu / rho
      nu_cst = nu * 1e6  # centistokes (cSt) 변환

      return {
         "Temperature (K)": T_kelvin,
         "Pressure (Pa)": P_pascal,
         "Specific Heat (Cp) [kJ/kg·K]": cp / 1000,
         "Specific Volume (v) [m^3/kg]": v,
         "Kinematic Viscosity (ν) [cSt]": nu_cst,
         "Prandtl Number (Pr)": pr,
         "Thermal Conductivity (k) [W/m·K]": k,
         "Density (ρ) [kg/m^3]": rho
      }

   except Exception as e:
      return f"Error occurred: {e}\n(Tip: REFPROP 설치 경로 확인 또는 성분비 합이 1.0인지 확인하세요.)"

# 원본 코드 그대로    
@app.route("/api/calc", methods=["POST"])
def calc():
    res = {}
    inputs = request.get_json()
    Cold_Q = inputs.get('tube_fluid_quantity_total')
    La_Heat = inputs.get('La_Heat')
    tube_ea = inputs.get('tube_no')
    tube_L = inputs.get('tube_OD')
    Hot_Q = inputs.get('shell_fluid_quantity_total')
    d = inputs.get('tube_pitch')
    Hot_intemp1 = inputs.get('shell_temp_out_expected')
    Cold_Spec = inputs.get('tube_temp_out_expected')
    Temp_in = inputs.get('tube_temp_in')
    Tk = inputs.get('Thk')
    Fouling_F = inputs.get('fouling_resistance')
    A = inputs.get('title')
    L = inputs.get('no_passes_per_shell_tube')
    M = inputs.get('no_passes_per_shell_shell')
    Hot_T = inputs.get('shell_temp_in')
    Hot_spec = inputs.get('shell_spec_temp')
    Shell_per_unit = inputs.get('shell_per_unit')
    Inlet_pressure_hot = inputs.get('inlet_pressure_shell')
    Inlet_pressure_cold = inputs.get('inlet_pressure_tube')

    Inlet_pressure_cold_Pa = (Inlet_pressure_cold + 1) * 1e5
    Inlet_pressure_hot_Pa = (Inlet_pressure_hot + 1) * 1e5

    t_pitch = d
    N = 160
    Temp_low = -10
    # 온도입력
    H_1 = 0
    H_2 = 30
    H_3 = 60
    H_4 = 90
    H_5 = 120
    H_6 = 150
    H_7 = 180

    P_stat = 50000

    VO1 = 1 / CP.PropsSI('D', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    VO2 = 1 / CP.PropsSI('D', 'T', H_2 + 273.15, 'P', P_stat, 'Nitrogen')
    VO3 = 1 / CP.PropsSI('D', 'T', H_3 + 273.15, 'P', P_stat, 'Nitrogen')
    VO4 = 1 / CP.PropsSI('D', 'T', H_4 + 273.15, 'P', P_stat, 'Nitrogen')
    VO5 = 1 / CP.PropsSI('D', 'T', H_5 + 273.15, 'P', P_stat, 'Nitrogen')
    VO6 = 1 / CP.PropsSI('D', 'T', H_6 + 273.15, 'P', P_stat, 'Nitrogen')
    VO7 = 1 / CP.PropsSI('D', 'T', H_7 + 273.15, 'P', P_stat, 'Nitrogen')
    VO_IN  = VO1

    C_CP1 = CP.PropsSI('CPMASS', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    CP2 = CP.PropsSI('CPMASS', 'T', H_2 + 273.15, 'P', P_stat, 'Nitrogen')
    CP3 = CP.PropsSI('CPMASS', 'T', H_3 + 273.15, 'P', P_stat, 'Nitrogen')
    CP4 = CP.PropsSI('CPMASS', 'T', H_4 + 273.15, 'P', P_stat, 'Nitrogen')
    CP5 = CP.PropsSI('CPMASS', 'T', H_5 + 273.15, 'P', P_stat, 'Nitrogen') 
    CP6 = CP.PropsSI('CPMASS', 'T', H_6 + 273.15, 'P', P_stat, 'Nitrogen') 
    CP7 = CP.PropsSI('CPMASS', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    CP_IN = C_CP1

    C_TC1 = CP.PropsSI('L', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    TC2 = CP.PropsSI('L', 'T', H_2 + 273.15, 'P', P_stat, 'Nitrogen')
    TC3 = CP.PropsSI('L', 'T', H_3 + 273.15, 'P', P_stat, 'Nitrogen')
    TC4 = CP.PropsSI('L', 'T', H_4 + 273.15, 'P', P_stat, 'Nitrogen')
    TC5 = CP.PropsSI('L', 'T', H_5 + 273.15, 'P', P_stat, 'Nitrogen')
    TC6 = CP.PropsSI('L', 'T', H_6 + 273.15, 'P', P_stat, 'Nitrogen')
    TC7 = CP.PropsSI('L', 'T', H_7 + 273.15, 'P', P_stat, 'Nitrogen')
    TC_IN = C_TC1
    
    C_PR1 = CP.PropsSI('PRANDTL', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    PR2 = CP.PropsSI('PRANDTL', 'T', H_2 + 273.15, 'P', P_stat, 'Nitrogen')
    PR3 = CP.PropsSI('PRANDTL', 'T', H_3 + 273.15, 'P', P_stat, 'Nitrogen')
    PR4 = CP.PropsSI('PRANDTL', 'T', H_4 + 273.15, 'P', P_stat, 'Nitrogen')
    PR5 = CP.PropsSI('PRANDTL', 'T', H_5 + 273.15, 'P', P_stat, 'Nitrogen')
    PR6 = CP.PropsSI('PRANDTL', 'T', H_6 + 273.15, 'P', P_stat, 'Nitrogen')
    PR7 = CP.PropsSI('PRANDTL', 'T', H_7 + 273.15, 'P', P_stat, 'Nitrogen')
    PR_IN = C_PR1

    C_VI1 = CP.PropsSI('V', 'T', H_1 + 273.15, 'P', P_stat, 'Nitrogen')
    VI2 = CP.PropsSI('V', 'T', H_2 + 273.15, 'P', P_stat, 'Nitrogen')
    VI3 = CP.PropsSI('V', 'T', H_3 + 273.15, 'P', P_stat, 'Nitrogen')
    VI4 = CP.PropsSI('V', 'T', H_4 + 273.15, 'P', P_stat, 'Nitrogen')
    VI5 = CP.PropsSI('V', 'T', H_5 + 273.15, 'P', P_stat, 'Nitrogen')
    VI6 = CP.PropsSI('V', 'T', H_6 + 273.15, 'P', P_stat, 'Nitrogen')
    VI7 = CP.PropsSI('V', 'T', H_7 + 273.15, 'P', P_stat, 'Nitrogen')
    VI_IN = C_VI1

    Req = Cold_Q * La_Heat
    Q = Req
    Tu_t = Tk*2   #튜브 두께 

    QWCP = .81
    TCL = .86
    DET = .0002  # 수렴

    Hot = Hot_Q/3600.
    Hot_V = Hot//0.029/943
    Hot_V = Hot//0.023/943
    res['shell_velocity'] = Hot_V
    H_TC = 0.35*1.16   # w/m-k
    H_TC = 0.35   # kcal

    TOC = 0  # Total_Convec
    TCP = 0  # Specific heat

    T1 = Temp_in
    T2 = Hot_intemp1  
    pai = np.pi

    water_VO = 1 / 1000

    # Water Liquid COLD Liquid
    C_VO = -1.38931902e-13* T1**4 + 3.39458511e-11 *T1**3 + 7.23194774e-10 * T1**2 + 1.87586756e-07* T1 + 9.97692559e-04
    C_CP = 1.0
    C_TC =(-1.17187500e-07* T1**4 + 5.14467593e-05*T1**3 +-1.49288194e-02*T1**2 + 2.28990741e+00*T1 + 5.57883333e+02)/1000*TCL
    C_VI = (7.81250000e-11* T1**4 +-3.07638889e-08*T1**3 + 4.81041667e-06*T1**2 +-3.81575397e-04*T1 + 1.59400000e-02)/10000
    C_PR =  7.91666667e-08* T1**4 +-2.98101852e-05*T1**3 + 4.39784722e-03*T1**2 +-3.22542791e-01*T1 + 1.19163333e+01

    res['specific_gravity_in_tube'] = water_VO / C_VO
    res['viscosity_in_tube '] = C_VI * 1000 / C_VO
    res['specific_heat_in_tube'] = C_CP * 4184
    res['thermal_conductivity_in_tube'] = C_TC * 4184

    # Glycol Water / 5 barG / 20~80
    H_VO=1/(-9.72222222225747E-10*T1**6 +2.95833333334354E-07*T1**5 +-0.000036180555555674*T1**4 +0.00226875000000702*T1**3 +-0.0788722222224456*T1**2 +0.985166666670283*T1 +1064.59999999998)    #VO 단위질량당부피 [m3/kg]
    H_CP=(9.72222222237005E-14*T1**6 +-2.87500000004451E-11*T1**5 +3.28472222227577E-09*T1**4 +-1.53958333336613E-07*T1**3 +-2.02944444433701E-06*T1**2 +0.00117368333333155*T1 +0.751150000000012)    #CP 비열 [kcal/kgK]
    H_TC=(-8.35960638194554E-13*T1**6 +2.4720550300897E-10*T1**5 +-2.93183338109684E-08*T1**4 +1.77522212668052E-06*T1**3 +-6.02584312600458E-05*T1**2 +0.00169876755517174*T1 +0.315219260533114)   #Thermal Conduct 열전도도 [kcal/mK]
    GW_cP=(-1.90277777778148E-11*T1**6 +4.38750000001133E-09*T1**5 +-2.82152777779164E-07*T1**4 +-1.03604166665801E-05*T1**3 +0.00251590555555266*T1**2 +-0.166062833333284*T1 +5.55809999999967)   #Viscosity 점도 [cP]
    H_VI=(-9.44444444448553E-18*T1**6 +1.7166666666791E-15*T1**5 +4.72222222071443E-15*T1**4 +-2.45249999999067E-11*T1**3 +2.73962222221913E-09*T1**2 +-1.57384666666615E-07*T1 +5.15039999999966E-06)   #Kinetic Viscosity 동점성계수 [m2/s]
    H_PR=(-5.63150143643295E-11*T1**6 +6.98315976359687E-09*T1**5 +1.06381819268007E-06*T1**4 +-0.000287155332857528*T1**3 +0.0273229496066831*T1**2 +-1.48490105059018*T1 +46.6024012145047)   #Prantl number 

    res['specific_gravity_in_shell'] = water_VO / H_VO
    res['viscosity_in_shell'] = H_VI / H_VO
    res['specific_heat_in_shell'] = H_CP * 4184
    res['thermal_conductivity_in_shell'] = H_TC * 4184

    C_CP1 = C_CP 
    C_PR1 = C_PR
    C_VI1 = C_VI
    C_TC1 = C_TC
    VO1 = C_VO
    SG1 = 1/C_VO/1000

    H_CP1 = H_CP
    H_TC1 = H_TC
    H_VO1 = H_VO
    GW_cP1 = GW_cP
    GW_PR1 = H_PR
    H_VI1 = H_VI
    Hd = t_pitch

    Surface_area = pai * tube_L**2 / 4  * tube_ea

    Cold_V = Cold_Q * C_VO / 3600 / (tube_ea * pai * (d-Tu_t)**2/ 4)
    Volume_Rate1 = Cold_Q*C_VO

    res['tube_velocity'] = Cold_V

    Cold_intemp = Temp_in
    Outlet_Temp = Cold_intemp

    Hot_intemp_LMTD = Hot_intemp1
    Cold_Re = Cold_V * d / C_VI
    Nu = 0.023 * Cold_Re**.8 * C_PR**0.4

    Cold_h = Nu * C_TC / d   
    Hot_Red = Hot_V * d / 0.0000016
    Hot_Nu = 0.664 * Hot_Red**.5 * H_PR**.3333
    Hot_h = Hot_Nu * H_TC / t_pitch 

    Total_convec = 1/(1/Hot_h + 1/Cold_h + 0.0005 )
    Total_convec = 110

    surface_a = Q / (Total_convec*(Hot_intemp1-Cold_intemp))
    surface_a = Q / (Total_convec)
    surface_a1= surface_a

    Hot_outtemp =  Hot_intemp1 - Q/(Hot_Q*H_CP) 
    BGW_outtemp = Hot_outtemp 
    TQ1 = Q
    GW_intemp = Q/(Hot_Q*H_CP)

    T_surface = 0
    TOQ = Q
    Hot_intemp = Hot_outtemp

    N = int(N)    # 구간분활 갯수 
    BLNG_intemp = Temp_in

    DPT = 0
    DpgT = 0
    DP_in = 0
    DPL = DP_in 

    Del_T = (Hot_intemp - Cold_intemp)

    R = 1/Hot_h + 1/Cold_h 
    SKT1 = Hot_intemp - Del_T * (1/Hot_h)/R
    SKT_out = Hot_intemp - Del_T * (1/Hot_h)/R
    Minim = 30
    Hd = t_pitch
    Total_convec1 = Total_convec
    SKT_out = Temp_in
    CPP = C_CP
    C_CP1 = C_CP 

    fig = plt.figure(figsize=(13,8))
    plt.axis ([Surface_area*-.15, Surface_area*1.15, Temp_low, Hot_T+20])

    start_T = time.time()
    Int_T = time.time()

    # 변수 선언 관련 에러 해결
    TG1 = 0
    Lc = 0
    Total_convec_0 = 0
    Total_convec3 = 0
    ToL_Spec = 0
    ToL_Q = 0
    Int_Hot_intemp = 0
    Int_Cold_intemp = 0
    T_surface2 = .001
    Int_Cold_intemp2 = 0
    Cold_outtemp = 0
    SKT = 0
    DP_out = 0

    res['specific_gravity_out_tube'] = ''
    res['viscosity_out_tube'] = ''
    res['specific_heat_out_tube'] = ''
    res['thermal_conductivity_out_tube'] = ''
    res['specific_gravity_out_shell'] = ''
    res['viscosity_out_shell'] = ''
    res['specific_heat_out_shell'] = ''
    res['thermal_conductivity_out_shell'] = ''


    for k in range (0, 20, 1) :   # GW 출구 Temp  예상
        TT2 =  Hot_intemp1 + k / 10
        print ("_________  stage", k , " - ", f"{TT2 :.1f} "" -  " f"{ Int_T-start_T :.1f} sec", " - ", round(Hot_outtemp+k/10,1) ,"___________") 
        T_surface = 1
        c1 = 0
        Total_convec = 650
        DPT = 0
        DpgT = 0
        Hot_intemp = TT2
        Cold_intemp = Temp_in 
        TOQ = 0
        TOC = 0
        SKT_out = Temp_in

        for i  in range (min(N, 360 - 2 * N), max(N, 360 - 2 * N), 1) :
            ii = i/2
            c1 = Total_convec 
            CPP1 = C_CP  
            Total_convec_0 = Total_convec            
            surface_a = 2*tube_ea* math.sin(math.radians(ii))*tube_L/2 * (math.cos(math.radians(ii))*tube_L/2- math.cos(math.radians(ii+1/2))*tube_L/2)
            if surface_a < 0:   
                print(surface_a, ii)
            T_surface =  T_surface + surface_a

            #  print ("---------------    stage", k," - ", i, " - ",  round(T2), "   ---------------")

            P_hot = Inlet_pressure_hot_Pa - DPT
            P_cold = Inlet_pressure_cold_Pa - DpgT

            # Water Liquid COLD Liquid
            T1 = Cold_intemp    
            C_VO = -1.38931902e-13* T1**4 + 3.39458511e-11 *T1**3 + 7.23194774e-10 * T1**2 + 1.87586756e-07* T1 + 9.97692559e-04
            C_CP = 1.0
            C_TC =(-1.17187500e-07* T1**4 + 5.14467593e-05*T1**3 +-1.49288194e-02*T1**2 + 2.28990741e+00*T1 + 5.57883333e+02)/1000*TCL
            C_VI = (7.81250000e-11* T1**4 +-3.07638889e-08*T1**3 + 4.81041667e-06*T1**2 +-3.81575397e-04*T1 + 1.59400000e-02)/10000
            C_PR =  7.91666667e-08* T1**4 +-2.98101852e-05*T1**3 + 4.39784722e-03*T1**2 +-3.22542791e-01*T1 + 1.19163333e+01
            C_cP = C_VI * 1000 / C_VO      # Centi-Poise

            res['specific_gravity_out_tube'] = water_VO / C_VO
            res['viscosity_out_tube'] = C_cP
            res['specific_heat_out_tube'] = C_CP * 4184
            res['thermal_conductivity_out_tube'] = C_TC * 4184

            T2 = Hot_intemp       
            x =  [H_1,     H_2,    H_3,    H_4,    H_5,    H_6,    H_7] 
            y =  [VO_IN,   VO2,    VO3,    VO4,    VO5,    VO6,    VO7 ] # 비체적 VO 3bar LN2

            def H_VO(x, y, T2) :  
                poly = np.polyfit(x, y, 6)
                H_VO = 1/(math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6]) 
                # H_VO = 1/((T2**6)*poly[0]+  (T2**5)*poly[1]+  (T2**4)*poly[2]+  (T2**3)*poly[3]+  (T2**2)*poly[4]+  (T2)*poly[5]+  poly[6])  # 46.3       
                # H_VO= (1/(np.power(T2,6)*poly[0] + np.power(T2,5)*poly[1]+ np.power(T2,4)*poly[2]+ np.power(T2,3)*poly[3]+np.power(T2,2)*poly[4]+ np.power(T2,1)*poly[5]+ poly[6])) 
                return H_VO
            H_VO = H_VO(x, y, T2)

            y = [CP_IN ,CP2, CP3, CP4, CP5, CP6, CP7] # Cp 
            def H_CP(x, y, T2) :  
                poly = np.polyfit(x, y, 6)
                H_CP = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])    
                return H_CP
            H_CP = H_CP(x,y,T2)

            y = [TC_IN, TC2, TC3, TC4, TC5, TC6, TC7] # TC 
            def H_TC(x, y, T2) :  
                poly = np.polyfit(x, y, 6)   
                H_TC = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+ math.pow(T2,1)*poly[5]+poly[6])*.86
                return H_TC
            H_TC = H_TC(x, y, T2) 

            y = [PR_IN, PR2, PR3, PR4, PR5, PR6, PR7] # PR 
            def H_PR(x, y, T2) : 
                poly = np.polyfit(x, y, 6)      
                H_PR =(math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])
                return H_PR
            H_PR = H_PR(x, y, T2)        

            y = [VI_IN, VI2, VI3, VI4, VI5, VI6, VI7] # VI 
            def H_VI(x, y, T2) :
                poly = np.polyfit(x, y, 6)
                H_VI = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])/10000
                return H_VI
            H_VI = H_VI(x,y,T2) 

            H_cP = H_VI * 1000 / H_VO # Centi-Poise 

            res['specific_gravity_out_shell'] = water_VO / H_VO
            res['viscosity_out_shell'] = H_cP
            res['specific_heat_out_shell'] = H_CP * 4184
            res['thermal_conductivity_out_shell'] = H_TC * 4184

            #  print(T2)
            #  print ("C_CP    :",round(C_CP,3), "  C_PR :",round(C_PR,2), "   C_VI :",round(C_VI,8),"    c_TC :",round(C_TC,3), "    C_VO :", round(C_VO,5))
            #  print("H_CP :", round(H_CP,3), "  H_PR :",round(H_PR,2)," H_VI :", round(H_VI,8)," H_TC :", round( H_TC,3), "  H_VO :", round(H_VO,3),
            #                "  H_cP :", round(H_cP,4) )
            #  input()
            Cold_V1 = Cold_Q / 3600           # 물
            Hd = t_pitch
            Area = tube_ea / 2 / L * Hd * math.sin(math.radians(ii)) * tube_L 
            # GW 유로 단면적(324mm dia, 324/2 = 0.16m)  tube_ea/2(글리콜 및 LNG유로) L = pass ea ,  Hd 유로간극

            Cold_V = Cold_V1/(Area * 1/C_VO)
            Cold_Re = Cold_V * Hd / C_VI

            Nu = 0.023 * math.pow(Cold_Re,.8) * math.pow(C_PR,0.333)
            if Cold_Re <  50000  :                                               # Tube 실 직경
                Nu = 0.664 * math.pow(Cold_Re,.5) * math.pow(C_PR,0.333)  
            
            Cold_h = Nu * C_TC / Hd                 
            Hot = Hot_Q  / 3600                                                      # 메탄
            Area = tube_ea / 2 / M * Hd * math.sin(math.radians(ii))  * tube_L       # GW 유로면적

            Hot_V = Hot / (Area * 1/H_VO)  
            Hot_Red = Hot_V * Hd / H_VI 

            Hot_Nu = 0.023 * math.pow(Hot_Red,.79) * math.pow(H_PR,.33)
            if  Hot_Red <  50000 : 
                Hot_Nu = 0.664 * math.pow(Hot_Red,.5) * math.pow(H_PR,.33)
            Hot_h = Hot_Nu * H_TC / Hd 

            if i == 158 : 
                Cold_VC =   Cold_V 
                Hot_VC = Hot_V
                C_cPC =  C_cP
                H_cPC = H_cP

            if i == 316 : 
                Cold_VCT =   Cold_V 
                Hot_VCT = Hot_V
                C_cPT =  C_cP
                H_cPT = H_cP

            Total_convec = 1/(1/Cold_h + 1/Hot_h + Fouling_F + Tk/14.4)        # 평행 평판 
        
            if i <= N :
                Total_convec_0 = 1/(1/Cold_h + 1/Hot_h + Fouling_F + Tk/14.4) * .85
                R = 1/Hot_h + 1/Cold_h + Fouling_F + Tk/14.4

            DP_in = 0

            for j in range(100000) :
                j = (j + 1)/150  
                TQ = (Total_convec)*surface_a*j

                Hot_outtemp =  Hot_intemp + TQ / (Hot_Q * H_CP)   # Counter Flow
                Cold_outtemp = Cold_intemp + TQ / (Cold_Q * C_CP) 
                
                hdt = Hot_intemp - Cold_outtemp           
                cdt = Hot_outtemp - Cold_intemp 

                if (hdt / cdt <= 0):
                    break

                LMTD = (hdt-cdt)/math.log(hdt/cdt)        
                TTQ = Total_convec * surface_a * LMTD
                DT = math.pow((TTQ - TQ), 2) /1000

                #  if N > N/2 :
                #   LMTD = (hdt-cdt)/math.log(hdt/cdt) * 1  # 온도 보정계수 0.9
                #   TTQ = Total_convec * surface_a * LMTD
                #   DT = math.pow((TTQ - TQ), 2) /1000    

                if DT <= DET :
                    Volume_Rate = Cold_Q * C_VO
                    TOQ = TOQ + TQ
                    ToL_Q = TOQ
                    ToL_W1 = ToL_Q  

                    #  Dp = (.0316*LNG_Re**-.25 * (tube_L + 100*Hd*4)         / N) * (1/VO) * (Velocity*L)**2 /(2 * Hd) # glycol side(수력학적직경)
                    Dpg = (64/Cold_Re * (tube_L + 100*Hd)                   / N) * (1/C_VO) * (Cold_V*L)**2 /(2 * Hd) # glycol side(수력학적직경)
                    # H.dp
                    # 패닝식임. 상기식은 밸브 300, 확대관 60등 총괄 확대축소 밸브등을 전부 합친 값임. 30% margin
                    # 브라지우스식 난류 f = .316 x Re^-.25, 3000<Re<100000
                    # HHI 실험결과를 참조하였슴.
                    #  Dp =  (4*.02 * (tube_L + 32*(d-Tu_t)*8) / N) * (1/VO) * Velocity**2 /(2 * (d-Tu_t)) # 20% margin
                    #  Dp1 =  1.5*(1/VO)*Velocity**2/2       

                #  Dpg = (0.05 * (tube_L + Hd*1000)                 / N) * (1/5.45) *  (GW_V*L)**2 / (2 * Hd) # glycol side(수력학적직경)
                    Dp = ( 64/Hot_Red* (tube_L + 100*Hd) / N) * (1/H_VO) * (Hot_V*L)**2 / (2 * Hd) # glycol side(수력학적직경)

                    DPT = DPT + Dp      # 유체 속도 및 비중량이 길이 방향에 따라 변하기 때문에 길이방향을 N개(60~120)로 미분하여 계산하고 적분함.
                    DpgT = DpgT + Dpg
                    DP_in = (DPT - Dp) / 10000
                    DP_out = DPT / 10000
                        
                    Del_T = (Hot_outtemp - Cold_outtemp)
                    SKT = Cold_outtemp + Del_T * Hot_h/(Cold_h + Hot_h)
                    SKT_in = SKT_out           
                    SKT_out =  SKT   

                    TOC = TOC + Total_convec
                    TCP = TCP + C_CP   

                    Location = 0                         # 비열
                    DTT = Location                       # Start Scale 변경, 비열입구 온도 위치 

                    CPP1 = CPP1*100 - C_CP1*100 + Location
                    CPP2 = C_CP*100 - C_CP1*100 + Location

                    Lc = 140                             #  열전달계수 입구위치 
                    Scailc = Total_convec1 - Lc

                    c1 = c1/6  - Scailc
                    c2 = Total_convec/6  - Scailc       
                
                    #  print ("Acumulate Area  :" , round(T_surface,3),"m2") 
                    #  print("Stage Heat Coef.(Kcal/m2.hr) :",round(Total_convec,1), "        Red  :",round(LNG_Re,1),GW_Red)
                    #  print("Stage Heat Capa.(Kcal/h)     :",round(TQ/1000,1), "          LMTD :",round(LMTD,2))
                    #  print("Cold in :",round(LNG_intemp,1), "         Cold out :",round(LNG_outtemp,1)) 
                    #  print("GW_in   :",round(Hot_intemp,1), "         GW out   :",round(GW_outtemp,1))
                    #  print("Acumulate Capa. (kcal/h) :",round(ToL_Q/1000,1),"(",round(ToL_Q/859.8,1),"KW )")
                    #  print("Cold Velocity :",round(Velocity,2),"m/s","   Press Loss : ", round(DpgT/1000,2),"Kpa") 
                    #  print("GW_V          :",round(GW_V,2),"m/s","    Press Loss : ", round(DPT/1000,2),"Kpa")         
                            
                    x = np.arange (0,400,1)
                    x = T_surface 
                    y = Cold_intemp
                    y1 = Cold_outtemp 
                    z = Hot_intemp
                    z1 = Hot_outtemp

                    a = DP_in * 20
                    a1 = DP_out * 20
                    s1 = SKT_in
                    s2 = SKT_out

                    line_Tk = .3             
                    plt.plot ([T_surface-surface_a, T_surface ],[y, y1 ] , c = 'b', lw = line_Tk+.4) 
                    plt.plot ([T_surface-surface_a, T_surface ],[z, z1 ] , c = 'r', lw = line_Tk+.5) 
                    if k == 6 : 
                        plt.plot ([T_surface-surface_a, T_surface ],[z,  z1 ], c = 'k', lw = line_Tk+.6) 

                    plt.plot ([T_surface-surface_a, T_surface ],[a, a1 ],  c = 'firebrick', lw = line_Tk)    # DP_in Temp 

                    plt.plot ([T_surface-surface_a, T_surface ],[s1, s2 ], c = 'g', lw = line_Tk)    # SKin Temp
                    plt.plot ([T_surface-surface_a, T_surface ],[c1, c2 ], c = 'm', lw = line_Tk)    # Convection
                    plt.plot ([T_surface-surface_a, T_surface ],[CPP1, CPP2 ], c = 'firebrick', lw = .5)   # Specific heat 

                    Cold_intemp = Cold_outtemp
                    Hot_intemp = Hot_outtemp 
                    DP_in = DP_out
                 
                    if (Hot_spec - Hot_outtemp) < .0005 :  # Hot 출구 예상온도 46도
                        Int_Cold_intemp =  Cold_intemp  
                        Int_Hot_intemp = Hot_outtemp
                        T_surface2 = T_surface 
                        Total_convec3 = Total_convec 
                        DPOUT1 =  DP_out

                    if (Cold_Spec - Cold_intemp) < .0001 :     # LNG 출구 예상온도 42도  
                        Int_Cold_intemp2 =  Cold_intemp  
                        Int_Hot_intemp2 = Hot_intemp     
                        T_surface2 = T_surface
                        Margine = (Surface_area - T_surface2)*100/ Surface_area
                        Margine = (Surface_area - T_surface)/ T_surface
                        ToL_W1 = ToL_Q / 859.8
                        DPOUT2 =  DP_out 
                        ToL_Spec = ToL_W1  
                        DPOUT2 =  DP_out
                        break
                            
                    if k==1 :
                        Q1 = ToL_Q/859.8
                        TG1 = Hot_outtemp  
                    #  print("T2:",TG2,"Q2:",Q2)  
                    #  plt.text(T_surface*1.02, Hot_outtemp - 10, round(Hot_outtemp,1), c = 'r')        # 100% 출구온도
                    if k==2 : 
                        Q2 = ToL_Q/859.8 
                        TG2 = Hot_outtemp         
                    if k==3 :
                        Q3 = ToL_Q/859.8 
                        TG3 = Hot_outtemp  
                    if k==4 :
                        Q4 = ToL_Q/859.8  
                        TG4 = Hot_outtemp   
                    if k==5 :
                        Q5 = ToL_Q/859.8  
                        TG5 = Hot_outtemp     
                    if k==6 :
                        Q6 = ToL_Q/859.8  
                        TG6 = Hot_outtemp    
                    if k==7 :
                        Q7 = ToL_Q/859.8  
                        TG7 = Hot_outtemp      
                    Int_T = time.time()
                    break   
        if (Hot_outtemp > Hot_T ) : break

    plt.rc('font', size=12)   
    plt.text(T_surface*.996, Hot_outtemp - 22, round(TG1,1), c = 'r')        # 100% 출구온도
    TOC = TOC / (320 - 2 * N + 1)
    TCP = TCP / (320 - 2 * N + 1) 
    #print(TT2) # GW Inlet Spec.

    plt.plot ([0],[surface_a1] , c = 'b',           label='COLD'         )            
    plt.plot ([0],[surface_a1] , c = 'r',           label='HOT')
    plt.plot ([0],[surface_a1] , c = 'firebrick',   label='P. Loss'  )
    plt.plot ([0],[surface_a1] , c = 'm',           label='Kcal/m2hr'   ) 

    plt.rc('font', size=12)
    plt.legend(loc=(.3,.02), ncol=1, frameon=True, shadow=True)
    plt.tick_params(axis='both', direction='in', length=6, pad=7, labelsize=8)
    plt.scatter (T_surface*1.05, Hot_intemp_LMTD*1.1 , c = 'r', alpha = .01)  # 그리드 크기 조정

    plt.rc('font', size=20)
    plt.xlabel('Surface Area (m^2)',fontsize=13)
    plt.ylabel('Temperature C',     fontsize=13) 

    plt.title(A,fontsize=13) 
    plt.rc('font', size=12) 
    plt.text(T_surface*-.1,Temp_in-10,'Cold_In', c = 'b')
    plt.text(T_surface*-.1,Temp_in-5,  round(Temp_in,1), c = 'b'  )                           # Cold 입구온도

    Sd = 6
    # plt.text(T_surface*-0.1, (CPP - CP1)*100 + Location - Sd,    'Specific_h',  c = 'firebrick')    # Heat transfer Coef. - -
    # plt.text(T_surface*-0.1, (CPP - CP1)*100 + Location + 0,      round(CP1,2), c = 'firebrick') 
    # plt.text(T_surface*1.02, (CP  - CP1)*100 + Location + Sd,    'Specific_h',  c = 'firebrick')
    # plt.text(T_surface*1.02, (CP  - CP1)*100 + Location + 0,      round(CP,2),  c = 'firebrick') 
    # plt.text(T_surface*.5,   (CP3 - CP1)*100 + Location + 0.,     round(CP3,2), c = 'firebrick')    # 50% 면적온도

    # Lc = 43
    DTS = Total_convec1 - Lc 
    Total_convec2 = Total_convec_0

    plt.rc('font', size=12) 
    plt.text(T_surface*-0.1, (Total_convec2 - DTS)/8 + 30,  'Kcal/m2h',              c = 'm')   # Heat transfer Coef.  - -
    plt.text(T_surface*-0.1, (Total_convec2 - DTS)/8 + 36,   round(Total_convec2),   c = 'm') 
    plt.text(T_surface*1.02, (Total_convec  - DTS)/8 + 30,  'Kcal/m2h',              c = 'm')
    plt.text(T_surface*1.02, (Total_convec -  DTS)/8 + 36,   round(Total_convec),    c = 'm') 
    plt.text(T_surface*.6,   (Total_convec3 - DTS)/8 + 40,   round(Total_convec3),   c = 'm')
    # plt.text(T_surface*.5,   (Total_convec3 - DTS)/8 + 36,   round(Total_convec3), c = 'g')   # 50% 면적온도 

    # plt.text (0, -56, 'Skin Temp. & Icing Cond. -36.C,(EG 50%)')
    # plt.text (0, -70, 'Minimum Skin Temp. :')
    # plt.text (7.2, -70, round(Minim,1) )
    # plt.plot ([0, T_surface ],[-36, -36 ] , c = 'g', lw = 1)

    plt.text(T_surface*-.1,Hot_intemp_LMTD-1, round(Hot_intemp_LMTD,1), c = 'r' )  # GW 입구온도
    plt.text(T_surface*.002,  TT2+2.5, round(TT2,1), c = 'r' )  # GW 입구온도
    plt.text(T_surface*-.1, -5, ' Press Loss', c = 'k' )                           # Press Loss
    plt.text(T_surface*-0.1,Hot_intemp_LMTD+5,'Hot_out', c = 'r')

    Hot_intemp_LMTD = Hot_intemp  # Pararell
    # Hot_intemp_LMTD = Hot_intemp1  # Count - flow
    plt.rc('font', size=14)
    aa = 1.2
    COR = 50
    # print((Hot_intemp_LMTD-Temp_in)/aa + COR)
    plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/aa + COR, 'Cold ')     
    plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/aa + COR, round(Cold_Q), c = 'b') 
    plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/aa + COR, 'Kg/h')
    bb = aa*1.15
    plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/bb+ COR, 'HOT ') 
    plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/bb+ COR, round(Hot_Q), c = 'r') 
    plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/bb+ COR, 'Kg/h ')
    cc = bb*1.18
    plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Spec ') 
    plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/cc+ COR, round(ToL_Spec),   c = 'k')
    plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Kw ')

    plt.text(T_surface*.5,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Toq :')
    plt.text(T_surface*.57,(Hot_intemp_LMTD-Temp_in)/cc+ COR, round(ToL_Q/860),  c = 'k')

    dd = cc*1.21
    plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/dd+ COR, 'C.dp')     
    plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/dd+ COR, round(DpgT/1000,1), c = 'b') 
    plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/dd+ COR, 'Kpa')
    ee = dd*1.28
    plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/ee+ COR, 'H.dp')     
    plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/ee+ COR, round(DPT/1000,1), c = 'firebrick')  
    plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/ee+ COR, 'Kpa')
    ff = ee*1.4
    plt.text(T_surface*.2, (Hot_intemp_LMTD-Temp_in)/ff+ COR, 'Fouling F')  
    plt.text(T_surface*.35,(Hot_intemp_LMTD-Temp_in)/ff+ COR, round(Fouling_F,7), c = 'm') 

    plt.rc('font', size=12)
    plt.text(Surface_area*.2, Int_Hot_intemp+3,  round(Int_Hot_intemp,1), c = 'r')  # 80% 면적온도
    plt.text(Surface_area*.2, Int_Cold_intemp-5, round(Int_Cold_intemp,1), c = 'b')

    # plt.text(Surface_area*.8, Int_LNG_intemp-6,'Surf. Area 80%' )
    # plt.text(T_surface2,  Int_Hot_intemp2-10, round(Int_Hot_intemp2,1), c = 'r')  # 80% Hot Temp
    plt.text(T_surface2,  Int_Cold_intemp2-8 , round(Int_Cold_intemp2,1), c = 'b')
    # plt.text(T_surface3,  Int_Hot_intemp-5, round(Int_Hot_intemp,1), c = 'r')
    # plt.text(T_surface*.6, DPOUT1-8, round(DPOUT1*10), c = 'k' )

    plt.text(T_surface*1.02, Hot_outtemp + 0,  round(Hot_outtemp,1), c = 'r')        # 100% 출구온도
    plt.text(T_surface*1.02, Hot_outtemp + 6,  'Hot_In' , c = 'r')
    plt.text(T_surface*1.02, Cold_outtemp - 2, round(Cold_outtemp,1), c = 'b')
    plt.text(T_surface*1.02, Cold_outtemp + 4, 'Cold_Out' , c = 'b')
    plt.text(T_surface*1.02, SKT + 4,          round(SKT,1) , c = 'g')
    plt.text(T_surface*1.02, SKT + 10,         'Skin_Temp' , c = 'g')

    # plt.text(T_surface, LNG_outtemp+2, '100%')
    # plt.text(T_surface2, DPOUT2*10, round(DPOUT2,1), c = 'r')  
    plt.text(T_surface*1.02, DP_out*10, round(DPT/1000,1), c = 'firebrick')
    plt.text(T_surface*1.07, DP_out*10, ' KPa',           c = 'firebrick')

    plt.rc('font', size=12)

    surface_pe = T_surface2*100/T_surface                                # 출구 온도 Spec.
    # plt.text (T_surface2, Int_LNG_intemp2+2,round(Int_LNG_intemp2,1,), c = 'b')    # 스펙온도에 따른 전열면적
    # plt.text(T_surface2+.5, Int_LNG_intemp2+2, '     %')  
    # plt.text(T_surface2, Int_LNG_intemp2-4,round(ToL_W1/1000,1)) 
    # plt.text(T_surface2+5, Int_LNG_intemp2-4,'KW')

    plt.gca().set_facecolor('oldlace')
    plt.rc ('font', family = "consolas", weight='bold', size=10)
    plt.text(T_surface*.88, Temp_low+10,'Coded by KIM & Mytec')

    plt.grid(True)
    print()
    print("----------------- TOTAL ---------------------")

    ToTal_K = TOC  # Average Heat Transfer Coef.
    ToL_Q = TOQ
    ToL_W = ToL_Q*1.163
    # ToTal_K = TOC  # Average Heat Transfer Coef.
    ToTal_K_W = ToTal_K * 1.163 
    print('surface_pe', surface_pe)
    margine = (100 - surface_pe) / surface_pe * 100
    mardumi = 1/(1 + margine/100)
    service = ToTal_K_W * mardumi
    clean = ToTal_K_W
    print('Margine', margine)
    print('Mardumi', mardumi)
    actual = 1 / (1 / clean + Fouling_F)
    surf_eff = T_surface * (actual / clean)
    LMTD = ToL_Spec *1000 / service / T_surface  
    LMTD = ToL_W / service / (T_surface * .99)

    print("Heat Capacity (kcal/h) :",round(ToL_Q/1000,1),"(",round(ToL_Q/859.8,1),"KW )")
    print("Surf. Area ( m^2 )  :" , round(T_surface,2))
    print("Heat Transfer Coef. (Kcal/m2.hr) :",round(ToTal_K,1),"(",round(ToTal_K_W,1),"W/m2.K)" )
    print("Transfer Service (W/m2.K) :",round(service,1) )
    print("Spec. Heat Capa  :", round(ToL_Spec,1), "KW")
    print("Fouling Factor :",round(Fouling_F,5))
    print()
    print("ToTal LMTD   :",round(LMTD,2))
    print("Cold In Temp :",round(Temp_in,1),"           Cold out Temp :",round(Cold_outtemp,1))
    print("HOT IN Temp  :",round(Hot_outtemp,1),"       HOT out Temp   :",round(Hot_intemp1,1))
    print("Hot_V        :",round(Hot_VC,1),"m/s","      Press Loss    :", round(DPT/1000,2),"Kpa")
    print() 
    # margine = (100 - surface_pe) / surface_pe * 100
    print("Sface_margin :",round(margine, 1),"%")
    print("Cold In_V.   :",round(Cold_V,1),"m/s",  "        Cold Out_V    :",round(Cold_VC,1),"m/s") 
    print("Cold_Q       :",round(Cold_Q),"Kg/h", "      HOT_Q          :", round(Hot_Q),"Kg/h")
    print("Tube_ea      :",round(tube_ea),"ea","         Tube Dia.     :" , round(d*1000,2),"mm")
    print()
    print("Cold In - Out property")
    print("Spec.H In  :",round(C_CP1,2), "   PR :",round(C_PR1,2), "   VI :",round(C_VI1,7),"     TC :",round(C_TC1,3), "   VO :", round(VO1,3))
    print("       Out :",round(C_CP,3),  "   PR :",round(C_PR,2),  "   VI :",round(C_VI,8), "   TC :",round(C_TC,3),  "   VO :", round(C_VO,3))
    print()
    print("HOT In - Out property")
    print("HOT_CP In  :", round(H_CP1,3),  "   HOT_VI :", round(H_VI1,8),"  HOT_TC :", round(H_TC1,3), "  HOT_VO :", round(H_VO1,3))
    print("       Out :", round(H_CP,3),   "       VI :", round(H_VI,8), "      TC :", round(H_TC,3),  "      VO :", round(H_VO,3))
    print()
    print ("Vol_Rate in(m3/h) :",round(Volume_Rate1,1),"  Vol_Rate out(m3/h) :",round(Volume_Rate,1))
    print("------------------------------------------------------")

    # print("Heat Capacity (KW) :",round(Q1,1),",  ",round(Q2,1),",  ",round(Q3,1),",  ",round(Q4,1),",  ",round(Q5,1))
    # print("HOT In             :",round(TG1,1),",  ",round(TG2,1),",  ",round(TG3,1),",  ",round(TG4,1),",  ",round(TG5,1))

    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)
    img_base64 = base64.b64encode(buf.read()).decode("utf-8")

    res['shell_side_out_temp'] = Hot_intemp1
    res['Tube Side Out Temp'] = Cold_outtemp
    res['mtd_corrected'] = LMTD
    res['shell_pressure_drop'] = DPT / 1000
    res['tube_pressure_drop'] = DpgT/1000
    res['transfer_rate_service'] = service
    res['transfer_rate_clean'] = clean
    res['transfer_rate_actual'] = actual
    res['surf_unit_gross'] = T_surface
    res['surf_unit_eff'] = surf_eff
    res['heat_exchanged'] = ToL_Q/859.8
    res['surf_shell_gross'] = T_surface * Shell_per_unit
    res['surf_shell_eff'] = surf_eff * Shell_per_unit
    # res['Tube No'] = tube_ea
    # res['Length'] = tube_L
    # res['OD'] = d
    # res['Thk (Avg)'] = Tk
    # res['Shell Side In Temp'] = Hot_intemp_LMTD
    # res['Tube Side In Temp'] = Temp_in
    # res['B9'] = Cold_Spec
    # res['B10'] = N
    # res['Fouling Factor'] = Fouling_F 
    # res['B12'] = La_Heat * 4.184
    # res['B13'] = ToL_Q / 859.8
    # res['B14'] = T_surface 
    # res['B17'] = Cold_Q
    # res['B18'] = Hot_Q
    # res['B19'] = Cold_VC 
    # res['B20'] = Cold_V
    # res['B21'] = C_TC1 * 1.163
    # res['B22'] = C_TC *1.163
    # res['B23'] = C_CP1 * 4.184
    # res['B24'] = C_CP  * 4.184
    # res['B25'] = ToTal_K																															
    # res['B26'] = SG1         # Gas in
    # res['B27'] = 1/C_VO/1000	# Gas	out																																											
    # res['B29'] = C_VI1 * 10000
    # res['B30'] = C_VI	* 10000																						
    # res['B31'] = Volume_Rate1 / 3600
    # res['B32'] = Volume_Rate	/ 3600
    # res['D1'] = A   # Project
    # res['D2'] = datetime.today().strftime("%Y. %m .%d")  # YYYY.mm.dd 형태의 출력																						
    # res['D3'] = B	# Custom	
    # res['D5'] = Hot_VC
    # res['D7'] = t_pitch
    # res['D8'] = 100-surface_pe
    # res['D12'] = L # Pass 
    # res['D14'] = C   # Type Shell&Tube, Plate&shell
    # res['D15'] = D   # LNG
    # res['D16'] = E   # GW 
    # res['D17'] = H_CP1 * 4.184 #
    # res['D18'] = H_CP * 4.184 # 
    # res['D19'] = H_TC1   #
    # res['D20'] = H_TC   #
    # res['D25'] = H_VO1   #
    # res['D26'] = H_VO   #
    # res['D27'] = Cold_VC   #
    # res['D28'] = Cold_VCT   #
    # res['D29'] = Hot_VC   #
    # res['D30'] = Hot_VCT   #
    # res['D31'] = C_cPC   #
    # res['D32'] = C_cPT   #
    # res['D33'] = H_cPC   #
    # res['D34'] = H_cPT   #
    # res['D35'] = TT2
    return jsonify({"resData" : res, "imgData" : img_base64})

# 다듬은 코드 - 결과값이 잘 안나와서 확인 필요
@app.route("/api/simulate", methods=["POST"])
def simulate():
   inputs = request.get_json()
   print("input type:", type(inputs))
   print("inputs:", inputs)
   res = {}
   
   hot_fluid_string = inputs.get('hot_fluid_string')
   cold_fluid_string = inputs.get('cold_fluid_string')
   # 입력값을 변수에 할당한다.
   cold_fluid_quantity = inputs.get('cold_fluid_quanitity') # LNG 유량 cold
   hot_fluid_quantity = inputs.get('hot_fluid_quantity') # GW 유량 ho
   tube_no = inputs.get('tube_no') # 튜브 개수
   tube_length = inputs.get('tube_length')

   # WARNING: 입력이 아니라 값으로. 어떤 값이 d인지 모르겠음
   d = .004   # 판 간격 .004 m 4mm
   M = 2
   L = 2

   hot_in_temp = inputs.get('hot_in_temp')
   cold_out_temp_input = inputs.get('cold_out_temp_input') # LNG 출구 spec.COLD
   cold_in_temp = inputs.get('cold_in_temp')
   avgerage_thickness = inputs.get('averageThickness') # 튜브 두께
   fouling_factor = inputs.get('foulingFactor')
   pitch = inputs.get('pitch') # plate pitch
   # M = inputs.get('M')
   # Hot_T = inputs.get('Hot_T')
   hot_out_temp_input = inputs.get('hot_out_temp_input')
   inlet_pressure_cold = inputs.get('inletPressureCold')
   inlet_pressure_hot = inputs.get('inletPressureHot')
   pressure_drop_allow_cold = inputs.get('pressureDropAllowCold')
   pressure_drop_allow_hot = inputs.get('pressureDropAllowHot')

   baffle_spacing = inputs.get('baffle_spacing')
   baffle_cut = inputs.get('baffle_cut')
   shell_inner_diameter = inputs.get('shell_inner_diameter')
   no_passes_per_shell_shell_side = inputs.get('no_passes_per_shell_shell_side')

   # barG => Pa로 변환
   inlet_pressure_cold = (inlet_pressure_cold + 1.01325) * 100000
   inlet_pressure_hot = (inlet_pressure_hot + 1.01325) * 100000
   
   # bar => Pa로 변횐
   pressure_drop_allow_cold = pressure_drop_allow_cold * 100000
   pressure_drop_allow_hot = pressure_drop_allow_hot * 100000

   # shell 관련 변수 mm => m로 단위 변환
   baffle_spacing = baffle_spacing / 1000
   shell_inner_diameter = shell_inner_diameter / 1000

   # 잠열 계산
   # WARNING : inlet pressure 기준으로 계산
   inlet_latent_heat_cold = CP.PropsSI("H", "P", inlet_pressure_cold, "Q", 1, cold_fluid_string) - CP.PropsSI("H", "P", inlet_pressure_cold, "Q", 0, cold_fluid_string)
   # inlet_latent_heat_hot =  CP.PropsSI("H", "P", inlet_pressure_hot, "Q", 1, hot_fluid_string) - CP.PropsSI("H", "P", inlet_pressure_hot, "Q", 0, hot_fluid_string)

   # 계산
   Req = cold_fluid_quantity * inlet_latent_heat_cold
   Q = Req
   Tu_t = avgerage_thickness * 2   #튜브 두께 

   DET = .0002  # 수렴

   cold_in_temp_kelvin = cold_in_temp + 273.15
   C_rho = CP.PropsSI('D', 'T', cold_in_temp_kelvin, 'P', inlet_pressure_cold, cold_fluid_string)
   C_VO = 1 / C_rho
   C_CP = CP.PropsSI('CPMASS', 'T', cold_in_temp_kelvin, 'P', inlet_pressure_cold, cold_fluid_string)
   C_TC = CP.PropsSI('L', 'T', cold_in_temp_kelvin, 'P', inlet_pressure_cold, cold_fluid_string)
   C_PR = CP.PropsSI('PRANDTL', 'T', cold_in_temp_kelvin, 'P', inlet_pressure_cold, cold_fluid_string)     
   C_VI = CP.PropsSI('V', 'T', cold_in_temp_kelvin, 'P', inlet_pressure_cold, cold_fluid_string)

   # Glycol Water / 5 barG / 20~80
   hot_in_temp_kelvin = hot_in_temp + 273.15
   H_rho = CP.PropsSI('D', 'T', hot_in_temp_kelvin, 'P', inlet_pressure_hot, hot_fluid_string)
   H_VO = 1 / H_rho
   H_CP = CP.PropsSI('CPMASS', 'T', hot_in_temp_kelvin, 'P', inlet_pressure_hot, hot_fluid_string)
   H_TC = CP.PropsSI('L', 'T', hot_in_temp_kelvin, 'P', inlet_pressure_hot, hot_fluid_string)
   H_PR = CP.PropsSI('PRANDTL', 'T', hot_in_temp_kelvin, 'P', inlet_pressure_hot, hot_fluid_string)      
   H_VI = CP.PropsSI('V', 'T', hot_in_temp_kelvin, 'P', inlet_pressure_hot, hot_fluid_string)

   Hd = pitch

   cold_velocity = cold_fluid_quantity * C_VO / 3600 / (tube_no * np.pi * (d - Tu_t)**2/ 4)
   cold_Re = cold_velocity * d / C_VI
   Nu = 0.023 * cold_Re**.8 * C_PR**0.4
   Cold_h = Nu * C_TC / d   

   hot_velocity = hot_fluid_quantity * H_VO / 3600 / no_passes_per_shell_shell_side / (baffle_cut * shell_inner_diameter * baffle_spacing)
   hot_Re = hot_velocity * d / 0.0000016
   Hot_Nu = 0.664 * hot_Re**.5 * H_PR**.3333
   Hot_h = Hot_Nu * H_TC / pitch 

   Total_convec = 1 / (1 / Hot_h + 1 / Cold_h + fouling_factor) # 열전달 계수

   surface_a = Q / (Total_convec*(hot_in_temp-cold_in_temp)) # 필요 열교환 면적

   hot_out_temp_local =  hot_in_temp - Q/(hot_fluid_quantity*H_CP) 

   res['Shell Side Velocity'] = hot_velocity
   res['Tube Side Velocity'] = cold_velocity

   # res['Boiling surface Area m^2'] = round(surface_a,2)
   # res['Heat Capa.(Kcal/h)'] = round(Q/1000,2)
   # res['Cold Inlet Temp'] = cold_in_temp
   # res['Cold Outlet Temp'] = cold_in_temp
   # res['GW Inlet Temp'] = hot_in_temp
   # res['GW Outlet Temp'] = round(hot_out_temp_local,2)

   total_surface = 0
   TOQ = Q
   hot_in_temp_local = hot_out_temp_local

   drop_pressure_hot = 0
   drop_pressure_cold = 0

   Hd = pitch

   # prevent unassociated error
   ToL_Spec = 0
   surface_result = 1

   for k in range (20) :   # GW 출구 Temp  예상
      hot_in_temp_local =  hot_in_temp
      cold_in_temp_local = cold_in_temp
      total_surface = 0 # 누적 표면적 합
      Total_convec = 650 # 전체 열전달 계수
      drop_pressure_hot = 0
      drop_pressure_cold = 0
      TOQ = 0
      TOC = 0
      TCP = 0

      # surface_a = 2*tube_ea* math.sin(math.radians(ii))*tube_L/2 * (math.cos(math.radians(ii))*tube_L/2- math.cos(math.radians(ii+1/2))*tube_L/2)   # 이게 무슨 코드야,,,
      # WARNING!! : 여기 수정 필요
      surface_a = tube_length * tube_no *  (avgerage_thickness / 2)
      total_surface =  total_surface + surface_a

      # drop pressure allow를 고려해서
      local_pressure_cold = max(inlet_pressure_cold - drop_pressure_cold, inlet_pressure_cold - pressure_drop_allow_cold)
      local_pressure_hot = max(inlet_pressure_hot - drop_pressure_hot, inlet_pressure_hot - pressure_drop_allow_hot)  

      # 국소 부위에서의 cold 물성치 계산
      cold_in_temp_local_kelvin = cold_in_temp_local + 273.15
      C_rho = CP.PropsSI('D', 'T', cold_in_temp_local_kelvin, 'P', local_pressure_cold, cold_fluid_string)
      C_VO = 1 / C_rho
      C_CP = CP.PropsSI('CPMASS', 'T', cold_in_temp_local_kelvin, 'P', local_pressure_cold, cold_fluid_string)
      C_TC = CP.PropsSI('L', 'T', cold_in_temp_local_kelvin, 'P', local_pressure_cold, cold_fluid_string)
      C_PR = CP.PropsSI('PRANDTL', 'T', cold_in_temp_local_kelvin, 'P', local_pressure_cold, cold_fluid_string)     
      C_VI = CP.PropsSI('V', 'T', cold_in_temp_local_kelvin, 'P', local_pressure_cold, cold_fluid_string)

      # 국소 부위에서의 hot 물성치 계산
      hot_in_temp_local_kelvin = hot_in_temp_local + 273.15
      H_rho = CP.PropsSI('D', 'T', hot_in_temp_local_kelvin, 'P', local_pressure_hot, hot_fluid_string)
      H_VO = 1 / H_rho
      H_CP = CP.PropsSI('CPMASS', 'T', hot_in_temp_local_kelvin, 'P', local_pressure_hot, hot_fluid_string)
      H_TC = CP.PropsSI('L', 'T', hot_in_temp_local_kelvin, 'P', local_pressure_hot, hot_fluid_string)
      H_PR = CP.PropsSI('PRANDTL', 'T', hot_in_temp_local_kelvin, 'P', local_pressure_hot, hot_fluid_string)      
      H_VI = CP.PropsSI('V', 'T', hot_in_temp_local_kelvin, 'P', local_pressure_hot, hot_fluid_string)

      fluid_quantity_cold_per_sec = cold_fluid_quantity / 3600
      Hd = pitch
      Area = tube_no / 2 / L * Hd * math.sin(math.radians(1)) * tube_length 

      # cold 흐름
      cold_velocity = fluid_quantity_cold_per_sec / (Area * 1/C_VO)
      cold_Re = cold_velocity * Hd / C_VI  # 레이놀즈 수

      Nu = 0.023 * math.pow(cold_Re,.8) * math.pow(C_PR,0.333)
      if cold_Re <  50000  :                                       
         Nu = 0.664 * math.pow(cold_Re,.5) * math.pow(C_PR,0.333)  
      
      Cold_h = Nu * C_TC / Hd        

      # hot 흐름
      fluid_quantity_hot_per_sec = hot_fluid_quantity  / 3600                                                      # 메탄
      Area = tube_no / 2 / M * Hd * tube_length       # GW 유로면적

      hot_velocity = fluid_quantity_hot_per_sec / (Area * 1/H_VO)  
      hot_Re = hot_velocity * Hd / H_VI 

      Hot_Nu = 0.023 * math.pow(hot_Re,.79) * math.pow(H_PR,.33)
      if  hot_Re <  50000 : 
         Hot_Nu = 0.664 * math.pow(hot_Re,.5) * math.pow(H_PR,.33)
      Hot_h = Hot_Nu * H_TC / Hd 

      Total_convec = 1 / (1 / Cold_h + 1 / Hot_h + fouling_factor + avgerage_thickness / 14.4)        # 평행 평판 

      for j in range(10) :
         print(j)
         delta_temp_local = hot_in_temp_local - cold_in_temp_local
         print(Total_convec, surface_a, delta_temp_local)
         TQ = (Total_convec) * surface_a * delta_temp_local # 국소 열전달량
         print('========', k, j, TQ, '========')

         hot_out_temp_local =  hot_in_temp_local - TQ / (hot_fluid_quantity * H_CP)   # Counter Flow
         cold_out_temp_local = cold_in_temp_local + TQ / (cold_fluid_quantity * C_CP) 
         print('========', hot_out_temp_local, cold_out_temp_local, '========')
         
         hdt = hot_in_temp_local - cold_out_temp_local           
         cdt = hot_out_temp_local - cold_in_temp_local 

         # counter flow에서 LMTD
         LMTD = (hdt-cdt)/math.log(hdt/cdt)     
         TTQ = Total_convec * surface_a * LMTD
         DT = math.pow((TTQ - TQ), 2) /1000
         print('DT', DT)

         #if DT <= DET :
         if True:
            TOQ = TOQ + TQ
            ToL_Q = TOQ
            ToL_W1 = ToL_Q  
            # warning: 여기 N은 임의로 결정함. 
            drop_pressure_cold_local = (64 / cold_Re * (tube_length + 100 * Hd) / 100) * (1 / C_VO) * (cold_velocity * L) ** 2 / (2 * Hd) # 국소 압력 강하, Cold
            drop_pressure_hot_local = (64 / hot_Re * (tube_length + 100 * Hd) / 100) * (1 / H_VO) * (hot_velocity * L) ** 2 / (2 * Hd) # 국소 압력 강하, Hot
            drop_pressure_hot = drop_pressure_hot + drop_pressure_hot_local # 누적 압력 강하, Hot
            drop_pressure_cold = drop_pressure_cold + drop_pressure_cold_local # 누적 압력 강하, Cold
            TOC = TOC + Total_convec
            TCP = TCP + C_CP   

            cold_in_temp_local = cold_out_temp_local
            hot_in_temp_local = hot_out_temp_local                  

            if math.pow ((hot_out_temp_input - hot_out_temp_local),2) < .0005 :
               surface_result = total_surface 

            if math.pow((cold_out_temp_input - cold_in_temp),2) < .0001 :
               surface_result = total_surface
               ToL_W1 = ToL_Q / 859.8
               ToL_Spec = ToL_W1  
               break  
      if (hot_out_temp_local > hot_in_temp ) : break

   surface_pe = round(surface_result*100/total_surface, 1)
   cold_outlet_temp = cold_out_temp_local
   hot_outlet_temp = hot_out_temp_local
   cold_outlet_pressure = inlet_pressure_cold - drop_pressure_cold
   hot_outlet_pressure = inlet_pressure_hot - drop_pressure_hot

   ## ==== 여기는 결과 출력 ==== ##
   ToTal_K = TOC  # Average Heat Transfer Coef.
   ToL_Q = TOQ
   ToL_W = ToL_Q*1.163
   ToTal_K_W = ToTal_K * 1.163 

   margine = (100 - surface_pe) / surface_pe * 100
   mardumi = 1/(1 + margine / 100)
   service = ToTal_K_W * mardumi
   LMTD = ToL_W / service / (total_surface * .99)

   # 열교환기 성능 관련 결과값들
   res['Heat Exchanged'] = round(ToL_Spec, 2)
   res['MTD (Corrected)'] = round(LMTD, 1)
   res['Transfer Rate, Service'] = round(service, 2)
   res['Transfer Rate, Clean'] = round(ToTal_K_W, 2)
   res['Transfer Rate, Actual'] = round(service * 0.99, 2)

   # 물성치 계산
   res['Tube Temperature (Out)'] = round(cold_outlet_temp, 2)
   res['Shell Temperature (Out)'] = round(hot_outlet_temp, 2)

   res['Cold Out Vapor'] = ''
   res['Cold Out Liquid'] = ''
   res['Cold Out Steam'] = ''
   res['Cold Out Water'] = ''
   res['Cold Out Noncondensables'] = ''
   res['Hot Out Vapor'] = ''
   res['Hot Out Liquid'] = ''
   res['Hot Out Steam'] = ''
   res['Hot Out Water'] = ''
   res['Hot Out Noncondensables'] = ''

   cold_out_phase = get_phase(cold_fluid_string, cold_outlet_temp, cold_outlet_pressure)
   hot_out_phase = get_phase(hot_fluid_string, hot_outlet_temp, hot_outlet_pressure)

   if cold_out_phase == 'Vapor':
      res['Cold Out Vapor'] = cold_fluid_quantity
   elif cold_out_phase == 'Liquid':
      res['Cold Out Liquid'] = cold_fluid_quantity
   elif cold_out_phase == 'Steam':
      res['Cold Out Steam'] = cold_fluid_quantity
   elif cold_out_phase == 'Water':
      res['Cold Out Water'] = cold_fluid_quantity
   else:
      res['Cold Out Noncondensables'] = cold_fluid_quantity

   if hot_out_phase == 'Vapor':
    res['Hot Out Vapor'] = hot_fluid_quantity
   elif hot_out_phase == 'Liquid':
      res['Hot Out Liquid'] = hot_fluid_quantity
   elif hot_out_phase == 'Steam':
      res['Hot Out Steam'] = hot_fluid_quantity
   elif hot_out_phase == 'Water':
      res['Hot Out Water'] = hot_fluid_quantity
   else:
      res['Hot Out Noncondensables'] = hot_fluid_quantity

   # Specific gravity
   rho_water_4C = CP.PropsSI("D", "T", 273.15 + 4, "P", 101325, "Water")
   rho_cold_in = CP.PropsSI("D", "T", cold_in_temp + 273.15, "P", inlet_pressure_cold, cold_fluid_string)
   rho_cold_out = CP.PropsSI("D", "T", cold_outlet_temp + 273.15, "P", cold_outlet_pressure, cold_fluid_string)
   rho_hot_in = CP.PropsSI("D", "T", hot_in_temp + 273.15, "P", inlet_pressure_hot, hot_fluid_string)
   rho_hot_out = CP.PropsSI("D", "T", hot_outlet_temp + 273.15, "P", hot_outlet_pressure, hot_fluid_string)

   res['Tube Specific Gravity In'] = round(rho_cold_in / rho_water_4C, 1)
   res['Tube Specific Gravity Out'] = round(rho_cold_out / rho_water_4C, 1)
   res['Shell Specific Gravity In'] = round(rho_hot_in / rho_water_4C, 1)
   res['Shell Specific Gravity Out'] = round(rho_hot_out / rho_water_4C, 1)

   # Viscosity
   viscosity_cold_in = CP.PropsSI("V", "T", cold_in_temp + 273.15, "P", inlet_pressure_cold, cold_fluid_string)
   viscosity_cold_out = CP.PropsSI("V", "T", cold_outlet_temp + 273.15, "P", cold_outlet_pressure, cold_fluid_string)
   viscosity_hot_in = CP.PropsSI("V", "T", hot_in_temp + 273.15, "P", inlet_pressure_hot, hot_fluid_string)
   viscosity_hot_out = CP.PropsSI("V", "T", hot_outlet_temp + 273.15, "P", hot_outlet_pressure, hot_fluid_string)

   res['Tube Viscosity In'] = round(viscosity_cold_in, 4)
   res['Tube Viscosity Out'] = round(viscosity_cold_out, 4)
   res['Shell Viscosity In'] = round(viscosity_hot_in, 4)
   res['Shell Viscosity Out'] = round(viscosity_hot_out, 4)

   # Specific Heat
   specific_heat_cold_in = CP.PropsSI("CPMASS", "T", cold_in_temp + 273.15, "P", inlet_pressure_cold, cold_fluid_string)
   specific_heat_cold_out = CP.PropsSI("CPMASS", "T", cold_outlet_temp + 273.15, "P", cold_outlet_pressure, cold_fluid_string)
   specific_heat_hot_in = CP.PropsSI("CPMASS", "T", hot_in_temp + 273.15, "P", inlet_pressure_hot, hot_fluid_string)
   specific_heat_hot_out = CP.PropsSI("CPMASS", "T", hot_outlet_temp + 273.15, "P", hot_outlet_pressure, hot_fluid_string)

   res['Tube Specific Heat In'] = round(specific_heat_cold_in, 1)
   res['Tube Specific Heat Out'] = round(specific_heat_cold_out, 1)
   res['Shell Specific Heat In'] = round(specific_heat_hot_in, 1)
   res['Shell Specific Heat Out'] = round(specific_heat_hot_out, 1)

   # Thermal Conductivity
   thermal_conductivity_cold_in = CP.PropsSI("L", "T", cold_in_temp + 273.15, "P", inlet_pressure_cold, cold_fluid_string)
   thermal_conductivity_cold_out = CP.PropsSI("L", "T", cold_outlet_temp + 273.15, "P", cold_outlet_pressure, cold_fluid_string)
   thermal_conductivity_hot_in = CP.PropsSI("L", "T", hot_in_temp + 273.15, "P", inlet_pressure_hot, hot_fluid_string)
   thermal_conductivity_hot_out = CP.PropsSI("L", "T", hot_outlet_temp + 273.15, "P", hot_outlet_pressure, hot_fluid_string)

   res['Tube Thermal Conductivity In'] = round(thermal_conductivity_cold_in, 4)
   res['Tube Thermal Conductivity Out'] = round(thermal_conductivity_cold_out, 4)
   res['Shell Thermal Conductivity In'] = round(thermal_conductivity_hot_in, 4)
   res['Shell Thermal Conductivity Out'] = round(thermal_conductivity_hot_out, 4)

   #outlet_latent_heat_cold = CP.PropsSI("H", "P", cold_outlet_pressure, "Q", 1, cold_fluid_string) - CP.PropsSI("H", "P", cold_outlet_pressure, "Q", 0, cold_fluid_string)
   #outlet_latent_heat_hot =  CP.PropsSI("H", "P", hot_outlet_pressure, "Q", 1, hot_fluid_string) - CP.PropsSI("H", "P", hot_outlet_pressure, "Q", 0, hot_fluid_string)

   res['Tube Latent Heat In'] = round(inlet_latent_heat_cold, 4)
   #res['Tube Latent Heat Out'] = round(outlet_latent_heat_cold, 4)
   # res['Shell Latent Heat In'] = round(inlet_latent_heat_hot, 4)
   #res['Shell Latent Heat Out'] = round(outlet_latent_heat_hot, 4)

   res['Tube Pressure Drop Calc'] = round(drop_pressure_cold / 1000, 3)
   res['Shell Pressure Drop Calc'] = round(drop_pressure_hot / 1000, 3)
   res['Shell Velocity'] = round(hot_velocity, 2)
   res['Tube Velocity'] = round(cold_velocity, 2)

   print(res)
   return jsonify({"printData" : res})

@app.route("/api/plot")
def plot():
    # 예시 그래프 데이터
    x = [1, 2, 3, 4]
    y = [10, 20, 25, 30]

    # 이미지 생성
    fig, ax = plt.subplots()
    ax.plot(x, y)
    ax.set_title("Sample Plot")
    ax.set_xlabel("X-axis")
    ax.set_ylabel("Y-axis")

    # 버퍼에 PNG 저장
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)

    # Base64 인코딩
    img_base64 = base64.b64encode(buf.read()).decode("utf-8")
    return jsonify({"image": img_base64})

@app.route("/api/excel", methods=["POST"])
def excel():
    inputs = request.get_json()
    heat_exchanged = inputs.get("heat_exchanged")
    mtd_corrected = inputs.get("mtd_corrected")
    shell_pressure_drop = inputs.get("shell_pressure_drop")
    shell_side_out_temp = inputs.get("shell_side_out_temp")
    shell_velocity = inputs.get("shell_velocity")
    specific_gravity_in_shell = inputs.get("specific_gravity_in_shell")
    specific_gravity_in_tube = inputs.get("specific_gravity_in_tube")
    specific_gravity_out_shell = inputs.get("specific_gravity_out_shell")
    specific_gravity_out_tube = inputs.get("specific_gravity_out_tube")
    specific_heat_in_shell = inputs.get("specific_heat_in_shell")
    specific_heat_in_tube = inputs.get("specific_heat_in_tube")
    specific_heat_out_shell = inputs.get("specific_heat_out_shell")
    specific_heat_out_tube = inputs.get("specific_heat_out_tube")
    surf_shell_gross = inputs.get("surf_shell_gross")
    surf_shell_eff = inputs.get("surf_shell_eff")
    surf_unit_eff = inputs.get("surf_unit_eff")
    surf_unit_gross = inputs.get("surf_unit_gross")
    thermal_conductivity_in_shell = inputs.get("thermal_conductivity_in_shell")
    thermal_conductivity_in_tube = inputs.get("thermal_conductivity_in_tube")
    thermal_conductivity_out_shell = inputs.get("thermal_conductivity_out_shell")
    thermal_conductivity_out_tube = inputs.get("thermal_conductivity_out_tube")
    transfer_rate_actual = inputs.get("transfer_rate_actual")
    transfer_rate_clean = inputs.get("transfer_rate_clean")
    transfer_rate_service = inputs.get("transfer_rate_service")
    tube_pressure_drop = inputs.get("tube_pressure_drop")
    tube_side_out_temp = inputs.get("tube_side_out_temp")
    tube_velocity = inputs.get("tube_velocity")
    viscosity_in_shell = inputs.get("viscosity_in_shell")
    viscosity_in_tube = inputs.get("viscosity_in_tube")
    viscosity_out_shell = inputs.get("viscosity_out_shell")
    viscosity_out_tube = inputs.get("viscosity_out_tube")   

    wb = load_workbook('output_tema.xlsx')
    ws = wb['1 - TEMA']

    ws["N32"] = heat_exchanged
    ws["BC32"] = mtd_corrected
    ws['AG30'] = shell_pressure_drop
    ws['AG20'] = shell_side_out_temp
    ws['U29'] = shell_velocity
    ws['U21'] = specific_gravity_in_shell
    ws['AS21'] = specific_gravity_in_tube
    ws['AG21'] = specific_gravity_out_shell
    ws['BE21'] = specific_gravity_out_tube
    ws['U25'] = specific_heat_in_shell
    ws['AS25'] = specific_heat_in_tube
    ws['AG25'] = specific_heat_out_shell
    ws['BE25'] = specific_heat_out_tube
    ws['BH10'] = surf_shell_eff
    ws['BB10'] = surf_shell_gross
    ws['Q10'] = surf_unit_eff
    ws['L10']= surf_unit_gross
    ws['U26'] = thermal_conductivity_in_shell
    ws['AS26'] = thermal_conductivity_in_tube
    ws['AG26'] = thermal_conductivity_out_shell
    ws['BE26'] = thermal_conductivity_out_tube
    ws['BC33'] = transfer_rate_actual
    ws['AI33'] = transfer_rate_clean
    ws['N33'] = transfer_rate_service
    ws['BE30'] = tube_pressure_drop
    ws['BE20'] = tube_side_out_temp
    ws['AS29'] = tube_velocity
    ws['U23'] = viscosity_in_shell
    ws['AS23'] = viscosity_in_tube
    ws['AG23'] = viscosity_out_shell
    ws['BE23'] = viscosity_out_tube

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name="result.xlsx",
        as_attachment=True
    )


def start_flask():
    app.run(debug=True)

if __name__ == "__main__":
    app.run(debug=True)
    #t = threading.Thread(target=start_flask, daemon=True)
    #t.start()
    #webview.create_window("React + Python", "http://127.0.0.1:5000")
    #webview.start()